


# ============================================================
#  Input handling
# ============================================================

import os
import sys
import openpyxl
from openpyxl.styles import Font
from PySide6.QtGui import QIcon  
from PySide6.QtWidgets import QHBoxLayout, QLabel, QFileDialog
from PySide6.QtWidgets import QTableWidget, QTableWidgetItem
from PySide6.QtWidgets import QGroupBox
from PySide6.QtWidgets import QMessageBox
from PySide6.QtWidgets import QCheckBox
from PySide6.QtWidgets import QToolButton
from PySide6.QtGui import QIcon
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QSlider


from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QFormLayout,
    QLineEdit, QComboBox, QTextEdit, QPushButton, QTabWidget, QLabel, 
    QDialog, QHBoxLayout, QSpinBox, QTableWidget, QTableWidgetItem, 
    QGroupBox, QHeaderView, QFileDialog, QMessageBox,QCheckBox  
)

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QCheckBox, QAbstractScrollArea, QMessageBox
)







# ============================================================
#  GLOBAL CONSTANTS
# ============================================================

# ------------------------------------------------------------
# Extra F1 Catalog (Alphabetical)
# ------------------------------------------------------------
EXTRA_F1_CATALOG = [
    ("ATA", "Grid transformer"),
    ("AXB", "Fire protection system"),
    ("BBT", "MV auxiliary transformer"),
    ("BDV", "Emergency diesel generator"),
    ("BFA", "Low voltage electrical supply"),
    ("BFT", "Auxiliary transformer"),
    ("CBA", "Energy management system"),
    ("CBB", "SCADA system"),
    ("CBC", "3rd level monitoring system"),
    ("CBD", "4th level monitoring system"),
    ("MSC", "Generator switch system"),
]


# ------------------------------------------------------------
# Extra F1 Templates (Prefix-Based Design)
# ------------------------------------------------------------
EXTRA_F1_TEMPLATES = {

    "CBC": {
        "description": "Monitoring System",
        "structure": {
            "UC": {"default": 1, "description": "Monitoring Panel"},
            "TF": {"default": 1, "description": "Network Switch"},
            "AG": {"default": 2, "description": "Industrial PC"},
        }
    },

    "CBD": {
        "description": "Metering System",
        "structure": {
            "UC": {"default": 1, "description": "Metering Panel"},
            "PG": {"default": 4, "description": "AC Meter"},
        }
    },

    # Add other systems here in alphabetical order...
}






# ============================================================
#  Parsing for Exception Handling
# ============================================================

def parse_mse_exceptions(text):
    result = {}
    for entry in text.strip().split(';'):
        if ':' in entry:
            g, vals = entry.strip().split(':')
            try:
                mse_count, tb, ta = map(int, vals.strip().split(','))
                result[g.strip()] = (mse_count, tb, ta)
            except:
                pass
    return result

def parse_mqa_exceptions(text):
    result = {}
    for entry in text.strip().split(';'):
        if ':' in entry:
            g, vals = entry.strip().split(':')
            try:
                mqa_count, wc, pv = map(int, vals.strip().split(','))
                result[g.strip()] = (mqa_count, wc, pv)
            except:
                pass
    return result



def parse_battery_exceptions(text):
    result = {}
    for entry in text.strip().split(';'):
        if ':' in entry:
            g, vals = entry.strip().split(':')
            try:
                unit, bank, rack = map(int, vals.strip().split(','))
                result[g.strip()] = (unit, bank, rack)
            except:
                pass
    return result



# ============================================================
#  Excel generator logic
# ============================================================


def generate_rdspp_excel(site_code, site_name, system_type, g0x_count, config):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gox"

    headers = ["Main Function F0", "System/Sub Systems F1", "Basic Function F2", "P1", "Description", "Field Label", "RDSPP Code"]
    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)

    ws.append([
        "", "", "", "",
        site_name,      # Description
        "",              # Field Label EMPTY
        site_code        # RDSPP Code
    ])

    for i in range(1, g0x_count + 1):
        g_code = f"G{str(i).zfill(2)}"
        ws.append([
            "", "", "", "",
            f"{site_name} Generator {str(i).zfill(2)}",   # Description
            "",                                           # Field Label
            f"{site_code}.{g_code}"                       # RDSPP Code
        ])



    for i in range(1, g0x_count + 1):
        g_code = f"G{str(i).zfill(2)}" 
        generator_number = int(g_code[-2:]) 
        generator_str = f"Generator {generator_number}"

        

        # ===============================
        # AHQ - Field Distribution System
        # ===============================

        ahq_config = config.get('AHQ_CONFIG', {})

        for main_code, sub_count in ahq_config.items():

            flabel = f"{g_code}.{main_code}"
            rdc = f"{site_code}.{g_code}.{main_code}"

            ws.append([
                g_code,
                main_code,
                "",
                "",
                f"Field Distribution System {main_code[-2:]} {generator_str}",
                flabel,
                rdc
            ])

            base_number = int(main_code[-2:])

            for j in range(1, sub_count + 1):

                sub_code = f"AHQ{base_number + j}"

                flabel = f"{g_code}.{sub_code}"
                rdc = f"{site_code}.{g_code}.{sub_code}"

                ws.append([
                    g_code,
                    sub_code,
                    "",
                    "",
                    f"Subsystem {str(j).zfill(2)} {main_code}",
                    flabel,
                    rdc
                ])




        # MST
        mst_code = "MST01"
        mst_flabel = f"{g_code}.{mst_code}"
        mst_rdc = f"{site_code}.{g_code}.{mst_code}"
        description = f"Transformer System 01 {generator_str}"
        ws.append([g_code, mst_code, "", "", description, mst_flabel, mst_rdc])

        for t in range(1, config['MST'] + 1):
            ta = f"TA{str(t).zfill(3)}"
            flabel = f"{g_code}.{mst_code}.{ta}"
            rdc = f"{site_code}.{g_code}.{mst_code}.{ta}"
            ws.append([g_code, mst_code, ta, "", f"Transformer Unit {t}", flabel, rdc])

        # MSE
        if g_code in config['mse_exceptions']:
            mse_count, tb_count, ta_count = config['mse_exceptions'][g_code]
        else:
            mse_count = config['MSE_DEFAULT']
            tb_count = config['TB']
            ta_count = config['TA']

        for p in range(1, mse_count + 1):
            mse_code = f"MSE{str(p).zfill(2)}"
            mse_flabel = f"{g_code}.{mse_code}"
            mse_rdc = f"{site_code}.{g_code}.{mse_code}"
            ws.append([g_code, mse_code, "", "", f"Inverter System {p} {generator_str}", mse_flabel, mse_rdc])

            for t in range(1, tb_count + 1):
                tb = f"TB{str(t).zfill(3)}"
                flabel = f"{g_code}.{mse_code}.{tb}"
                rdc = f"{site_code}.{g_code}.{mse_code}.{tb}"
                ws.append([g_code, mse_code, tb, "", f"Inverter Module {t} Inverter System {p}", flabel, rdc])

            for t in range(1, ta_count + 1):
                ta = f"TA{str(t).zfill(3)}"
                flabel = f"{g_code}.{mse_code}.{ta}"
                rdc = f"{site_code}.{g_code}.{mse_code}.{ta}"
                ws.append([g_code, mse_code, ta, "", f"DC-DC Converter {t} Inverter System {p}", flabel, rdc])

        # MQA
    
        for mqa_g, mqa_id, wc_count, pv_count in config.get("MQA_DETAIL", []):
            if mqa_g != g_code:
                continue  

            mqa_flabel = f"{g_code}.{mqa_id}"
            mqa_rdc = f"{site_code}.{g_code}.{mqa_id}"
            ws.append([g_code, mqa_id, "", "", f"PV Generator System {mqa_id} {generator_str}", mqa_flabel, mqa_rdc])

            # Combiner Boxes
            for c in range(1, wc_count + 1):
                wc = f"WC{str(c).zfill(3)}"
                flabel = f"{g_code}.{mqa_id}.{wc}"
                rdc = f"{site_code}.{g_code}.{mqa_id}.{wc}"
                ws.append([g_code, mqa_id, wc, "", f"Combiner Box for {mqa_id}", flabel, rdc])

                for w in range(1, config['WD_PER_WC'] + 1):
                    wd = f"WD{20 + w:03}"
                    flabel = f"{g_code}.{mqa_id}.{wc}-{wd}"
                    rdc = f"{site_code}.{g_code}.{mqa_id}.{wc}-{wd}"
                    ws.append([g_code, mqa_id, wc, wd, f"Core {w} LVDC Cable", flabel, rdc])

            # PV Strings
            for s in range(1, pv_count + 1):
                gc = f"GC{str(s).zfill(3)}"
                flabel = f"{g_code}.{mqa_id}.{gc}"
                rdc = f"{site_code}.{g_code}.{mqa_id}.{gc}"
                ws.append([g_code, mqa_id, gc, "", f"PV String {s}", flabel, rdc])

                for w in range(1, config['WD_PER_PV'] + 1):
                    wd = f"WD{10 + w:03}"
                    flabel = f"{g_code}.{mqa_id}.{gc}-{wd}"
                    rdc = f"{site_code}.{g_code}.{mqa_id}.{gc}-{wd}"
                    ws.append([g_code, mqa_id, gc, wd, f"Core {w} DC Cable", flabel, rdc])



        # Battery
        if g_code in config['BAT_EXCEPTIONS']:
            unit_count, bank_count, rack_count = config['BAT_EXCEPTIONS'][g_code]
        else:
            unit_count = config['BAT_UNIT']
            bank_count = config['BAT_BANK']
            rack_count = config['BAT_RACK']

        for u in range(1, unit_count + 1):
            batt_code = f"NQB{str(u).zfill(2)}"
            batt_flabel = f"{g_code}.{batt_code}"
            batt_rdc = f"{site_code}.{g_code}.{batt_code}"
            ws.append([g_code, batt_code, "", "", f"Battery Storage Unit {u} {generator_str}", batt_flabel, batt_rdc])

            for b in range(1, bank_count + 1):
                for r in range(1, rack_count + 1):
                    gb = f"GB{b}0{r}"
                    flabel = f"{g_code}.{batt_code}.{gb}"
                    rdc = f"{site_code}.{g_code}.{batt_code}.{gb}"
                    ws.append([g_code, batt_code, gb, "", f"Rack {r} Bank 0{b} Battery Storage Unit {u}", flabel, rdc])

        # ===============================
        # Extra F1 Systems
        # ===============================
        extra_config = config.get("EXTRA_F1_CONFIG", {})

        for f0_code, f0_data in extra_config.items():

            template = EXTRA_F1_TEMPLATES.get(f0_code, {})
            description = template.get("description", f0_code)

            for idx in range(1, f0_data["count"] + 1):

                f0_instance = f"{f0_code}{str(idx).zfill(2)}"

                flabel = f"{g_code}.{f0_instance}"
                rdc = f"{site_code}.{g_code}.{f0_instance}"

                ws.append([
                    g_code,
                    f0_instance,
                    "",
                    "",
                    f"{description} {str(idx).zfill(2)} {generator_str}",
                    flabel,
                    rdc
                ])

                # 🔥 Subsystems (no strict template dependency)
                for f1_code, f1_count in f0_data["structure"].items():

                    for j in range(1, f1_count + 1):

                        f1_instance = f"{f1_code}{str(j).zfill(3)}"

                        flabel = f"{g_code}.{f0_instance}.{f1_instance}"
                        rdc = f"{site_code}.{g_code}.{f0_instance}.{f1_instance}"

                        ws.append([
                            g_code,
                            f0_instance,
                            f1_instance,
                            "",
                            f"{f1_code} {str(j).zfill(2)}",
                            flabel,
                            rdc
                        ])


    return wb



# ============================================================
#  Y0X Generation Sheet
# ============================================================


import yaml

# Logical Part
def generate_y0x_sheet(wb, site_code, y0x_yaml_text):
    try:
        data = yaml.safe_load(y0x_yaml_text)
    except yaml.YAMLError as e:
        raise ValueError("Invalid YAML input for Y0x section")

    ws = wb.create_sheet("Y0x")  # This will create Sheet2

    headers = ["Main Function F0", "System/Sub Systems F1", "Basic Function F2", "P1", "Description", "Field Label", "RDSPP Code"]
    ws.append(headers)

    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)

    for f0, components in data.items():
        for entry in components:
            f1, f2, p1, desc = entry
            f1 = f1 or ""
            f2 = f2 or ""
            p1 = p1 or ""
            desc = desc or ""

            field_label = f"{f0}.{f1}"
            if f2:
                field_label += f".{f2}"
            if p1:
                field_label += f"-{p1}"

            rdcpp_code = f"{site_code}.{field_label}"

            row = [f0, f1, f2, p1, desc, field_label, rdcpp_code]
            ws.append(row)

# Dialogue box structure

class Y0xDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Y0x System Editor")
        self.resize(600, 400)

        layout = QVBoxLayout()
        self.y0x_input = QTextEdit()
        self.y0x_input.setPlaceholderText("Paste Y0x structure in YAML format here...")

        # Set default YAML template
        self.y0x_input.setPlainText("""                          
    The Standard structure for Yox is given in below format F0:-[F1,F2,P1, Description]. Based on requirements edit or delete.After editing delete this line. \n
Y01:
  - [AHA10, '', '', 'Distribution system 01']
  - [AHA11, '', '', 'Subsystem 01 Distribution system 01']
  - [AHA12, '', '', 'Subsystem 02 Distribution system 01']
  - [AHA13, '', '', 'Subsystem 03 Distribution system 01']

  - [CBA01, '', '', 'RWE Communication System']
  - [CBA01, UC001, '', 'RWE Communication Panel 01']
  - [CBA01, TF001, TF001, 'RWE Network Switch 01']
  - [CBA01, TF002, TF001, 'RWE Firewall']
  - [CBA01, TF002, TF002, 'RWE GreenBox Gateway']
  - [CBA01, TF003, TF001, 'RWE BT WAN Router']
  - [CBA01, TF003, TF002, 'RWE RUT956 4G/LTE Router']
                                    
  - [CBB01, '', '', 'PowerFactors Monitoring System']
  - [CBB01, UC001, '', 'PowerFactors Main SCADA Panel 01']
  - [CBB01, TF001, TF001, 'PowerFactors Core Switch 01']
  - [CBB01, TF002, TF001, 'PowerFactors Firewall']
  - [CBB01, KF002, KF001, 'PowerFactors RS485 Converter']
  - [CBB01, KF002, KF002, 'PowerFactors IO Module']
  - [CBB01, AG002, AG001, 'PV PPC']
  - [CBB01, AG002, AG002, 'BESS PPC']
  - [CBB01, AG002, AG003, 'Hybrid PPC']
  - [CBB01, AG001, AG001, 'PowerFactors SCADA Controller']
  - [CBB01, KF001, KF001, 'PowerFactors SCADA Server']
  - [CBB01, PH001, PH001, 'PowerFactors HMI']
                                    
  - [CBC01, '', '', 'RWE (Control Room) Monitoring System']
  - [CBC01, PG001, PG001, 'Tariff Meter']
  - [CBC01, PG001, PG002, 'Eberle PQM']
  - [CBC01, PG001, PG003, 'PPC PQM']
  - [CBC01, AG001, AG001, 'SEL RTAC']
                                    
  - [CBC02, '', '', 'RWE (33KV Switchroom) Monitoring System']
  - [CBC03, '', '', '132KV Compound Monitoring System']
                                    
  - [BBT01, '', '', 'MV Auxiliary Transformer System 01']
  - [BBT01, TA001, '', 'MV Auxiliary Transformer Unit 01']
                                    
  - [BFA01, '', '', 'LV electrical main supply system 01']
  - [BFA01, UC001, '', 'LV electrical main supply Panel 01']
                                    
  - [BMA01, '', '', 'Uninterruptable Power Supply System 01']
  - [BMA01, TB001, '', 'Uninterruptable Power Supply Unit 01']
                                    
  - [AXB01, '', '', 'Fire Protection System 01']
  - [AXB01, UC001, '', 'Fire Protection Panel 01']
""")

        btn_close = QPushButton("Close")
        btn_close.clicked.connect(self.accept)

        layout.addWidget(QLabel(" Enter Y0x Systems (YAML format):"))
        layout.addWidget(self.y0x_input)
        layout.addWidget(btn_close)

        self.setLayout(layout)

    def get_yaml(self):
        return self.y0x_input.toPlainText().strip()


# ============================================================
#  SAP Sheet generation
# ============================================================

def generate_sap_sheet(wb):
    ws_src = wb["Gox"]          
    ws_sap = wb.create_sheet("SAP")

    headers = [
        "Record Number",
        "FunctLocCategory",
        "Functional Location",
        "Structure indicator",
        "FunctLocDescrip.",
        "Language Key",
        "Object Type",
        "Manufacturer",
        "Mfr Ctry/Reg",
        "Model number",
        "Manuf. Serial Number",
        "Warranty",
        "Warranty end date",
        "ManufactPartNo.",
        "Planning Plant",
        "Catalog Profile",
        "Superior FunctLoc.",
        "Valid From",
        "Maintenance Plant",
        "Plant Section",
        "Location",
        "Room",
        "Main Work Center",
        "Plant for WorkCenter",
        "Maint. Planner Group",
        "Installation allowed",
        "User Statuses",
        "Inventory Number",
        "Sort Field",
        "System Statuses"
    ]

    ws_sap.append(headers)

    record_no = 1

    # Read from Gox sheet
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        description = row[4]    # Description column
        rdspp_code = row[6]     # RDSPP Code column

        if not rdspp_code:
            continue

        ws_sap.append([
            record_no,     # Record Number
            "P",           # FunctLocCategory
            rdspp_code,    # Functional Location
            "",            # Structure indicator
            description,   # FunctLocDescrip.
            "EN",          # Language Key
            "", "", "", "", "", "", "", "",
            "", "", "", "", "", "", "", "", "",
            "", "", "", "", "", ""
        ])

        record_no += 1


# ============================================================
#  Extra F1 Separate dialog box.
# ============================================================

class ExtraF1Dialog(QDialog):

    def __init__(self, f0_code, parent=None):
        super().__init__(parent)
        self.f0_code = f0_code
        self.setWindowTitle(f"{f0_code} Configuration")
        self.resize(400, 300)

        layout = QVBoxLayout()

        # Number of systems
        self.count_input = QLineEdit("1")

        # Subsystem table
        self.sub_table = QTableWidget()
        self.sub_table.setColumnCount(2)
        self.sub_table.setHorizontalHeaderLabels(["Subsystem Code", "Count"])
        self.sub_table.horizontalHeader().setStretchLastSection(True)

        btn_add = QPushButton("Add Subsystem")
        btn_add.clicked.connect(self.add_subsystem)

        btn_save = QPushButton("Save")
        btn_save.clicked.connect(self.accept)

        layout.addWidget(QLabel("Number of Systems"))
        layout.addWidget(self.count_input)
        layout.addWidget(self.sub_table)
        layout.addWidget(btn_add)
        layout.addWidget(btn_save)

        self.setLayout(layout)

        # Load default template
        self.load_template()


    def load_template(self):
        template = EXTRA_F1_TEMPLATES.get(self.f0_code)
        if not template:
            return

        for sub_code, sub_data in template["structure"].items():
            row = self.sub_table.rowCount()
            self.sub_table.insertRow(row)
            self.sub_table.setItem(row, 0, QTableWidgetItem(sub_code))
            self.sub_table.setItem(row, 1, QTableWidgetItem(str(sub_data["default"])))



    def add_subsystem(self):
        row = self.sub_table.rowCount()
        self.sub_table.insertRow(row)
        self.sub_table.setItem(row, 0, QTableWidgetItem("XX"))
        self.sub_table.setItem(row, 1, QTableWidgetItem("1"))


    def get_config(self):
        structure = {}

        for row in range(self.sub_table.rowCount()):
            code = self.sub_table.item(row, 0).text().strip().upper()
            count = int(self.sub_table.item(row, 1).text().strip())
            structure[code] = count

        return {
            "count": int(self.count_input.text()),
            "structure": structure
        }



# ============================================================
#  Application UI design
# ============================================================


class RDSPPApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("RDSPP Generator PV & BESS")
        self.setWindowIcon(QIcon("rdspp.ico"))
        self.resize(970,700)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.extra_f1_config = {}

        # Define default slider states before init_ui
        self.enable_wc_cables_flag = False
        self.enable_pv_cables_flag = False


        self.setStyleSheet("""
            QWidget {
                font-family: Segoe UI, sans-serif;
                font-size: 13px;
                background-color: #ffffff;
            }

            QLabel {
                font-weight: bold;
                margin-bottom: 4px;
                color: #333333;
            }

            QLineEdit, QComboBox, QTextEdit {
                padding: 6px;
                background-color: #ffffff;
                border: 1px solid #bbb;
                border-radius: 6px;
            }

            QPushButton {
                background-color: #6EC1E4;
                color: #ffffff;
                padding: 8px 16px;
                border: none;
                border-radius: 6px;
            }

            QPushButton:hover {
                background-color: #6EC1E4;
            }

            QCheckBox {
                spacing: 8px;
                padding: 4px;
            }


            QTabWidget::pane {
                border: 1px solid #ccc;
                border-radius: 4px;
                background-color: #ffffff;
            }

            QTabBar::tab {
                background-color: #e6f2f8;
                padding: 8px 14px;
                border: 1px solid #ccc;
                border-bottom: none;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
            }

            QTabBar::tab:selected {
                background-color: #ffffff;
                border-bottom: 1px solid white;
            }
        """)




        self.init_ui()



    def init_ui(self):
        layout = QVBoxLayout()


        # --- Project Information Group ---
        project_group = QGroupBox("Project Information")
        project_layout = QHBoxLayout()

        self.site_input = QLineEdit()
        self.site_input.setFixedWidth(300)

        self.site_name_input = QLineEdit()
        self.site_name_input.setFixedWidth(300)

        self.sys_type = QComboBox()
        self.sys_type.addItems(["Select", "PV", "PV + Battery"])
        self.sys_type.setFixedWidth(150)

        project_layout.addWidget(QLabel("Site Name:"))
        project_layout.addWidget(self.site_name_input)

        project_layout.addSpacing(20)

        project_layout.addWidget(QLabel("Site Code:"))
        project_layout.addWidget(self.site_input)

        project_layout.addSpacing(20)

        project_layout.addWidget(QLabel("System Type:"))
        project_layout.addWidget(self.sys_type)

        project_group.setLayout(project_layout)
        project_group.setMinimumHeight(100)  

        generator_group = QGroupBox("Main Function F0 Configuration")
        generator_layout = QHBoxLayout()




        # G0x input
        self.g0x_input = QLineEdit()
        self.g0x_input.setFixedWidth(100)


        # Y0x button and label
        self.open_y0x_btn = QPushButton("Click here")
        self.open_y0x_btn.setStyleSheet("padding: 6px;")
        self.open_y0x_btn.clicked.connect(self.open_y0x_dialog)

        self.y0x_label = QLabel(" Y0x System:")
        self.y0x_label.setStyleSheet("margin-left: 20px;")  # spacing between G0x and Y0x
        self.y0x_label.setVisible(True)

        generator_layout.addWidget(QLabel("Number of G0x:"))
        generator_layout.addWidget(self.g0x_input)

        generator_layout.addSpacing(20)

        generator_layout.addWidget(self.y0x_label)
        generator_layout.addWidget(self.open_y0x_btn)

        generator_layout.addStretch()
        generator_group.setLayout(generator_layout)




        # === Tabs for Components ===
        # === Grouped F1-Level Configuration Tabs ===
        f1_group = QGroupBox("F1-Level System Configuration")
        f1_layout = QVBoxLayout()

        self.tabs = QTabWidget()
        self.tabs.addTab(self.create_ahq_tab(), "🔌 AHQ_Field Distribution Systems")
        self.tabs.addTab(self.create_mst_tab(), "🔄 MST_Transformers")
        self.tabs.addTab(self.create_mse_tab(), "⚡ MSE_Inverters")

        self.tabs.addTab(self.create_mqa_tab(), "🔆 MQA_PV Generators")
        self.tabs.addTab(self.create_battery_tab(), "🔋 Battery")
        self.tabs.addTab(self.create_extra_f1_tab(), "📦 Extra F1 Systems")

        # Connect tab change signal
        self.tabs.currentChanged.connect(self.on_tab_changed)

        f1_layout.addWidget(self.tabs)
        f1_group.setLayout(f1_layout)


        

        # === Output ===
        output_box = QGroupBox("Output Settings")
        output_layout = QHBoxLayout()
        self.output_path_input = QLineEdit()
        browse_btn = QPushButton("Browse")
        browse_btn.clicked.connect(self.browse_output_path)
        output_layout.addWidget(QLabel("Output Path:"))
        output_layout.addWidget(self.output_path_input)
        output_layout.addWidget(browse_btn)
        output_box.setLayout(output_layout)

        
        # === Generate Button ===
        self.btn_generate = QPushButton("Generate RDSPP Excel")
        self.btn_generate.clicked.connect(self.generate_rdspp)

        # === Help Button ===
        self.btn_help = QPushButton("Help")
        self.btn_help.setFixedWidth(120)
        self.btn_help.clicked.connect(self.show_help)

        # === Final Layout ===
        layout.addWidget(project_group)
        layout.addWidget(generator_group)
        layout.addWidget(f1_group)

        # Help button (bottom-right of system tabs)
        help_layout = QHBoxLayout()
        help_layout.addStretch()
        help_layout.addWidget(self.btn_help)
        layout.addLayout(help_layout)

        # Output settings
        layout.addWidget(output_box)

        # Generate button
        layout.addWidget(self.btn_generate)

        self.central_widget.setLayout(layout)


    

    def open_y0x_dialog(self):
        dialog = Y0xDialog(self)
        if dialog.exec():
            self.y0x_yaml_text = dialog.get_yaml()
        else:
            self.y0x_yaml_text = ""  # Clear if dialog canceled



    def create_ahq_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        # Number of main AHQ systems
        self.ahq_main_count_input = QLineEdit("1")

        # Table for subsystems per main AHQ
        self.ahq_table = QTableWidget()
        self.ahq_table.setColumnCount(2)
        self.ahq_table.setHorizontalHeaderLabels(["Main AHQ Code", "Subsystem Count"])
        self.ahq_table.horizontalHeader().setStretchLastSection(True)

        btn_generate = QPushButton("Generate AHQ Main Systems")
        btn_generate.clicked.connect(self.generate_ahq_table)

        layout.addWidget(QLabel("Number of AHQ Main Systems"))
        layout.addWidget(self.ahq_main_count_input)

        layout.addWidget(btn_generate)

        layout.addWidget(QLabel("Configure Subsystems per AHQ"))
        layout.addWidget(self.ahq_table)

        tab.setLayout(layout)
        return tab


    def generate_ahq_table(self):
        try:
            count = int(self.ahq_main_count_input.text())
        except:
            QMessageBox.warning(self, "Input Error", "Enter valid AHQ count.")
            return

        self.ahq_table.setRowCount(0)

        for i in range(1, count + 1):
            base = i * 10
            main_code = f"AHQ{base}"

            row = self.ahq_table.rowCount()
            self.ahq_table.insertRow(row)

            self.ahq_table.setItem(row, 0, QTableWidgetItem(main_code))
            self.ahq_table.setItem(row, 1, QTableWidgetItem("3"))  # default subsystems


    def parse_ahq_config(self):
        result = {}

        for row in range(self.ahq_table.rowCount()):
            main_code = self.ahq_table.item(row, 0).text().strip()
            sub_count = int(self.ahq_table.item(row, 1).text().strip())

            result[main_code] = sub_count

        return result


    def create_mst_tab(self):
        self.mst_input = QLineEdit("1")
        tab = QWidget()
        layout = QFormLayout()
        layout.addRow("MST per G0x:", self.mst_input)
        tab.setLayout(layout)
        return tab

    def create_mse_tab(self):
        self.mse_count_input = QLineEdit("1")
        self.tb_input = QLineEdit("2")
        self.ta_input = QLineEdit("2")
        self.mse_exceptions = QLineEdit("G0x:F1,F21,F22; G0x:F1,F21,F22")
        # Tooltip on hover for MSE Exception input
        self.mse_exceptions.setToolTip(
            "🧠 <b>MSE Exceptions Format:</b><br>"
            "Use format -example <code>G01:2,4,3; G02:1,2,1</code><br>"
            "Here G01:2,4,3 → 2 MSEs, 4 Inverter Modules (TB), 3 DC-DC Converters (TA) for G01<br>"
            "<i>Use semicolons to separate multiple G0x entries</i>"
        )
        tab = QWidget()
        layout = QFormLayout()
        layout.addRow("MSE per G0x:", self.mse_count_input)
        layout.addRow("Inverter Modules(TB) per MSE:", self.tb_input)
        layout.addRow("DC-DC Converters(TA) per MSE:", self.ta_input)
        layout.addRow("MSE/TB/TA Exceptions for certain G0x:", self.mse_exceptions)
        tab.setLayout(layout)
        return tab

    def toggle_wc_cables(self):
        self.enable_wc_cables_flag = self.enable_wc_slider.value() == 1
        self.wc_status.setText("ON" if self.enable_wc_cables_flag else "OFF")
        self.wc_status.setStyleSheet(
            "font-weight: bold; color: green;" if self.enable_wc_cables_flag else "font-weight: bold; color: gray;"
        )

    def toggle_pv_cables(self):
        self.enable_pv_cables_flag = self.enable_pv_slider.value() == 1
        self.pv_status.setText("ON" if self.enable_pv_cables_flag else "OFF")
        self.pv_status.setStyleSheet(
            "font-weight: bold; color: green;" if self.enable_pv_cables_flag else "font-weight: bold; color: gray;"
        )

    
    def create_mqa_tab(self):
        tab = QWidget()
        main_layout = QVBoxLayout()

        #  Help button inside MQA tab
        help_btn = QToolButton()
        help_btn.setText("❓")
        help_btn.setToolTip("Click for help about MQA configuration")
        help_btn.setFixedSize(28, 28)
        help_btn.setStyleSheet("""
            QToolButton {
                background-color: #f0f8ff;
                border: 1px solid #87ceeb;
                border-radius: 4px;
                font-weight: bold;
                font-size: 16px;
            }
        """)
        help_btn.clicked.connect(lambda: QMessageBox.information(
            self,
            "💡 MQA Tab Help",
            "🔢 Edit number of MQA systems per G0x in the table 1\n"
            "🧮 Click 'Generate MQA Detail Table' then\n"
            "🛠 In table 2 edit Combiner box and PV String values for each MQA system as needed\n"
            "🔌 Slide cable core buttons to auto-generate 2 cable cores for each CB and PV strings"
        ))

        # === Top layout for help button ===
        top_layout = QHBoxLayout()
        top_layout.addWidget(help_btn)
        top_layout.addStretch()
        main_layout.addLayout(top_layout)

        # === Tables ===
        self.mqa_count_table = QTableWidget()
        self.mqa_count_table.setColumnCount(2)
        self.mqa_count_table.setHorizontalHeaderLabels(["G0x", "#MQA"])
        self.mqa_count_table.setFixedWidth(200)

        # Test row to start
        self.mqa_count_table.insertRow(0)
        self.mqa_count_table.setItem(0, 0, QTableWidgetItem("G01"))
        self.mqa_count_table.setItem(0, 1, QTableWidgetItem("2"))

        self.mqa_detail_table = QTableWidget()
        self.mqa_detail_table.setColumnCount(4)
        self.mqa_detail_table.setHorizontalHeaderLabels(["G0x", "MQA ID", "CombinerBox", "PV Strings"])
        self.mqa_detail_table.setFixedWidth(460)

        table_layout = QHBoxLayout()
        table_layout.addWidget(self.mqa_count_table)
        table_layout.addWidget(self.mqa_detail_table)
        main_layout.addLayout(table_layout)

        # Generate Button
        generate_detail_btn = QPushButton("Generate MQA Detail Table")
        generate_detail_btn.setFixedWidth(240)
        generate_detail_btn.setStyleSheet("font-weight: bold; background-color: #87ceeb; padding: 6px;")
        generate_detail_btn.clicked.connect(self.generate_mqa_detail_table)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(generate_detail_btn)
        btn_layout.addStretch()
        main_layout.addLayout(btn_layout)

        # === Sliders ===
        self.enable_wc_slider = QSlider(Qt.Horizontal)
        self.enable_wc_slider.setMinimum(0)
        self.enable_wc_slider.setMaximum(1)
        self.enable_wc_slider.setValue(0)
        self.enable_wc_slider.setFixedSize(40, 20)
        self.enable_wc_slider.setStyleSheet("""
            QSlider::groove:horizontal {
                height: 10px;
                background: lightgray;
                border-radius: 5px;
            }
            QSlider::handle:horizontal {
                width: 18px;
                background: white;
                border: 1px solid #555;
                border-radius: 9px;
                margin: -5px 0;
            }
        """)
        self.enable_wc_slider.valueChanged.connect(self.toggle_wc_cables)

        self.enable_pv_slider = QSlider(Qt.Horizontal)
        self.enable_pv_slider.setMinimum(0)
        self.enable_pv_slider.setMaximum(1)
        self.enable_pv_slider.setValue(0)
        self.enable_pv_slider.setFixedSize(40, 20)
        self.enable_pv_slider.setStyleSheet("""
            QSlider::groove:horizontal {
                height: 10px;
                background: lightgray;
                border-radius: 5px;
            }
            QSlider::handle:horizontal {
                width: 18px;
                background: white;
                border: 1px solid #555;
                border-radius: 9px;
                margin: -5px 0;
            }
        """)
        self.enable_pv_slider.valueChanged.connect(self.toggle_pv_cables)

        self.wc_status = QLabel("OFF")
        self.wc_status.setFixedWidth(30)
        self.wc_status.setStyleSheet("font-weight: bold; color: gray;")

        self.pv_status = QLabel("OFF")
        self.pv_status.setFixedWidth(30)
        self.pv_status.setStyleSheet("font-weight: bold; color: gray;")

        # === Cable Slider Layout with Descriptive Labels
        cable_layout = QHBoxLayout()
        cable_layout.addStretch()

        # 🔌 WC Cable
        wc_label = QLabel("WC Cable Cores")
        wc_label.setFixedWidth(100)
        wc_label.setStyleSheet("font-weight: bold;")
        cable_layout.addWidget(wc_label)
        cable_layout.addWidget(self.enable_wc_slider)
        cable_layout.addWidget(self.wc_status)

        cable_layout.addSpacing(40)

        # 🔋 PV Cable
        pv_label = QLabel("PV Cable Cores")
        pv_label.setFixedWidth(100)
        pv_label.setStyleSheet("font-weight: bold;")
        cable_layout.addWidget(pv_label)
        cable_layout.addWidget(self.enable_pv_slider)
        cable_layout.addWidget(self.pv_status)

        cable_layout.addStretch()
        main_layout.addLayout(cable_layout)

        # === Final assembly
        tab.setLayout(main_layout)
        return tab


    def parse_mqa_detail_table(self):
        mqa_details = []

        for row in range(self.mqa_detail_table.rowCount()):
            try:
                g_code = self.mqa_detail_table.item(row, 0).text().strip()
                mqa_id = self.mqa_detail_table.item(row, 1).text().strip()
                wc_count = int(self.mqa_detail_table.item(row, 2).text().strip())
                pv_count = int(self.mqa_detail_table.item(row, 3).text().strip())

                mqa_details.append((g_code, mqa_id, wc_count, pv_count))
            except Exception as e:
                QMessageBox.warning(self, "Parse Error", f"Error reading MQA row {row+1}: {e}")
                return []

        return mqa_details

    def on_tab_changed(self, index):
            tab_name = self.tabs.tabText(index)
            if "MQA" in tab_name:
                self.fill_mqa_count_table()


    def fill_mqa_count_table(self):
        try:
            g0x_text = self.g0x_input.text().strip()
            if not g0x_text.isdigit():
                return
            g0x_count = int(g0x_text)

        except ValueError:
            QMessageBox.warning(self, "Input Error", "Please enter a valid number of G0x systems.")
            return

        self.mqa_count_table.setRowCount(g0x_count)
        for i in range(g0x_count):
            g_code = f"G{str(i + 1).zfill(2)}"
            self.mqa_count_table.setItem(i, 0, QTableWidgetItem(g_code))
            self.mqa_count_table.setItem(i, 1, QTableWidgetItem("1"))  # Default 1 MQA per G0x

    def on_tab_changed(self, index):
        tab_name = self.tabs.tabText(index)
        if "MQA" in tab_name:
            self.fill_mqa_count_table()


    def generate_mqa_detail_table(self):
            
        self.mqa_detail_table.setRowCount(0)
        row = 0

        for i in range(self.mqa_count_table.rowCount()):
            g_code = self.mqa_count_table.item(i, 0).text().strip()
            try:
                mqa_count = int(self.mqa_count_table.item(i, 1).text().strip())
            except ValueError:
                QMessageBox.warning(self, "Input Error", f"Invalid MQA count for {g_code}")
                return

            for m in range(1, mqa_count + 1):
                mqa_id = f"MQA{str(m).zfill(2)}"
                self.mqa_detail_table.insertRow(row)
                self.mqa_detail_table.setItem(row, 0, QTableWidgetItem(g_code))
                self.mqa_detail_table.setItem(row, 1, QTableWidgetItem(mqa_id))
                self.mqa_detail_table.setItem(row, 2, QTableWidgetItem("2"))   # Default WC
                self.mqa_detail_table.setItem(row, 3, QTableWidgetItem("10"))  # Default PV
                row += 1


    def create_battery_tab(self):
        self.battery_unit_input = QLineEdit("1")
        self.battery_bank_input = QLineEdit("1")
        self.battery_rack_input = QLineEdit("5")
        self.battery_exceptions = QLineEdit("G0x:F1,F21,F22; G0x:F1,F21,F22")
        # Tooltip on hover for Battery Exception input
        self.battery_exceptions.setToolTip(
            "🧠 <b>Battery Exceptions Format:</b><br>"
            "Use format -example <code>G01:2,4,3; G02:1,2,1</code><br>"
            "Here G01:2,4,3 → 2 Battery, 4 banks, 3 racks<br>"
            "<i>Use semicolons to separate multiple G0x entries</i>"
        )
        tab = QWidget()
        layout = QFormLayout()
        layout.addRow("Battery Units per Gox:", self.battery_unit_input)
        layout.addRow("Banks per battery Unit :", self.battery_bank_input)
        layout.addRow("Racks per Bank:", self.battery_rack_input)
        layout.addRow("Battery/Banks/Racks exceptions per G0x:", self.battery_exceptions)
        tab.setLayout(layout)
        return tab
    

    def create_extra_f1_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        self.extra_f1_dropdown = QComboBox()
        self.extra_f1_dropdown.addItem("Select Extra F1 System", None)

        for code, name in EXTRA_F1_CATALOG:
            self.extra_f1_dropdown.addItem(f"{name} ({code})", code)

        btn_configure = QPushButton("Configure System")
        btn_configure.clicked.connect(self.open_extra_f1_dialog)

        # Summary table
        self.extra_f1_summary = QTableWidget()
        self.extra_f1_summary.setColumnCount(2)
        self.extra_f1_summary.setHorizontalHeaderLabels(["System", "Status"])
        self.extra_f1_summary.horizontalHeader().setStretchLastSection(True)

        layout.addWidget(QLabel("Select Extra F1 System"))
        layout.addWidget(self.extra_f1_dropdown)
        layout.addWidget(btn_configure)

        layout.addSpacing(15)
        layout.addWidget(QLabel("Configured Systems"))
        layout.addWidget(self.extra_f1_summary)

        tab.setLayout(layout)
        return tab

    
    def open_extra_f1_dialog(self):
        f0_code = self.extra_f1_dropdown.currentData()

        if not f0_code:
            QMessageBox.warning(self, "Select System", "Please select a system.")
            return

        dialog = ExtraF1Dialog(f0_code, self)

        if dialog.exec():
            config = dialog.get_config()

            self.extra_f1_config[f0_code] = config

            print("Saved Config:", self.extra_f1_config)

            self.update_extra_summary(f0_code)



    def update_extra_summary(self, f0_code):
        row = self.extra_f1_summary.rowCount()
        self.extra_f1_summary.insertRow(row)

        self.extra_f1_summary.setItem(row, 0, QTableWidgetItem(f0_code))
        self.extra_f1_summary.setItem(row, 1, QTableWidgetItem("Configured"))



    def create_y0x_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        self.y0x_input = QTextEdit()
        self.y0x_input.setPlaceholderText("Paste Y0x structure in YAML format here...")

        layout.addWidget(QLabel("📄 Enter Y0x Systems:"))
        layout.addWidget(self.y0x_input)

        tab.setLayout(layout)
        return tab


    def browse_output_path(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "RDSPP.xlsx", "Excel Files (*.xlsx)")
        if path:
            self.output_path_input.setText(path)

    def show_help(self):
        QMessageBox.information(
            self,
            "RDSPP Generator – Help",
            (
                " RDSPP Generator (PV & BESS)\n\n"
                "This tool helps you generate SAP-ready RDSPP and Functional Location "
                "Excel files for Solar (PV) and PV + Battery plants.\n\n"

                " How to use the tool:\n"
                "1. Enter the Site Code and Site Name.\n"
                "2. Select the Technology (PV or PV + Battery).\n"
                "3. Enter the number of Generators (G0x).\n"
                "4. Configure systems in each tab (AHQ, MST, MSE, MQA, Battery).Enter the number of components present for the site.\n"
                "5. Place the cursor in the exceptions tab for explanations and usage.\n"
                "6. The question mark help button in MQA tab provides explanations and usage.\n"
                "7. (Optional) Add Y0x systems using the Y0x editor.\n"
                "8. Choose an output path and click 'Generate RDSPP Excel'.\n\n"

                " Generated Excel sheets:\n"
                "• Gox  – RDSPP hierarchy for the plant\n"
                "• Y0x  – Optional auxiliary systems\n"
                "• SAP  – SAP Functional Location upload format\n\n"

                "Notes:\n"
                "• The tool does not modify SAP data directly.\n"
                "• Always review the Excel output before SAP upload.\n\n"

                "👨‍💻 Tool Developed by:\n"
                "Abhishek Sivasankara Pillai\n"
                "Work Student Data Management-Digital Operations-RWE Renewables(March 2024-March 2026)"
            )
        )


# ============================================================
# MAIN RDSPP GENERATION CONTROLLER
# ============================================================


    def generate_rdspp(self):
        print(" Step 1: Button Clicked......")

        site_base = self.site_input.text().strip()
        site_name = self.site_name_input.text().strip()
        tech = self.sys_type.currentText()

        if tech == "PV":
            site = f"{site_base}PV"
        elif tech == "PV + Battery":
            site = f"{site_base}PB"
        else:
            QMessageBox.warning(self, "Input Error", "Select valid technology")
            return


        g0x_text = self.g0x_input.text().strip()

        if not site or tech == "Select" or not g0x_text.isdigit():
            QMessageBox.warning(self, "Input Error", "Please provide valid inputs.")
            return

        g0x_count = int(g0x_text)


        config = {
            'AHQ_CONFIG': self.parse_ahq_config(),
            'MST': int(self.mst_input.text()),
            'MSE_DEFAULT': int(self.mse_count_input.text()),
            'TB': int(self.tb_input.text()),
            'TA': int(self.ta_input.text()),
            'mse_exceptions': parse_mse_exceptions(self.mse_exceptions.text()),
            'MQA_DEFAULT': 1,  # Default if no detail table is used
            'CB': 1,           # Set sensible defaults or connect to inputs if needed
            'STR': 10,
            'WD_PER_WC': 2 if self.enable_wc_cables_flag else 0,
            'WD_PER_PV': 2 if self.enable_pv_cables_flag else 0,
            'ENABLE_WC_CABLES': self.enable_wc_cables_flag,
            'ENABLE_PV_CABLES': self.enable_pv_cables_flag,

            'MQA_DETAIL': self.parse_mqa_detail_table(),
            'ENABLE_WC_CABLES': self.enable_wc_cables_flag,
            'ENABLE_PV_CABLES': self.enable_pv_cables_flag,

            'BAT_UNIT': int(self.battery_unit_input.text()),
            'BAT_BANK': int(self.battery_bank_input.text()),
            'BAT_RACK': int(self.battery_rack_input.text()),
            'BAT_EXCEPTIONS': parse_battery_exceptions(self.battery_exceptions.text()),
            'EXTRA_F1_CONFIG': getattr(self, "extra_f1_config", {}),
             
        }
        
        print("extra_f1_config BEFORE Excel:", getattr(self, "extra_f1_config", {}))
        output_path = self.output_path_input.text()
        if not output_path:
            output_path = os.path.join(os.getcwd(), f"RDSPP_{site}.xlsx")

        try:
            wb = generate_rdspp_excel(
                site,          # DE_NW.HAM01PB
                site_name,     # Hambach A
                tech,
                g0x_count,
                config
            )


            if self.y0x_yaml_text.strip():
                try:
                    generate_y0x_sheet(wb, site, self.y0x_yaml_text)
                except Exception as e:
                    QMessageBox.critical(self, "Y0x YAML Error", f"Error in Y0x section:\n{e}")
                    return

            generate_sap_sheet(wb)


            #  Save the final workbook with everything
            wb.save(output_path)
            print("Hello")
            #  Now show success message
            QMessageBox.information(self, "Success", f"Excel file saved to:\n{output_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error: {e}")

# ============================================================
# MAIN ENTRY
# ============================================================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = RDSPPApp()
    window.show()
    sys.exit(app.exec())